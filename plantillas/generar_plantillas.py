#!/usr/bin/env python3
"""Genera las plantillas del sitio.

Las planillas son la configuración del grupo: acá se define su estructura una
sola vez, para que sea igual en todos los grupos y se pueda regenerar.

    pip install openpyxl && python3 plantillas/generar_plantillas.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

AQUI = os.path.dirname(os.path.abspath(__file__))

NAVY, MENTA, CREMA, GRIS, AMBAR = '1E2A2E', 'D5F5EA', 'F9F4ED', '7F7F7F', 'FEF9E7'
BORDE = Border(bottom=Side(style='thin', color=NAVY))


def titulo(ws, texto, bajada, ancho):
    ws['A1'] = texto
    ws['A1'].font = Font(bold=True, size=14, color=NAVY)
    ws['A2'] = bajada
    ws['A2'].font = Font(italic=True, size=10, color=GRIS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho)


def encabezados(ws, cols, fila=3):
    for j, c in enumerate(cols, start=1):
        cel = ws.cell(row=fila, column=j, value=c)
        cel.font = Font(bold=True, color=NAVY, size=11)
        cel.fill = PatternFill('solid', fgColor=MENTA)
        cel.border = BORDE
    ws.freeze_panes = ws.cell(row=fila + 1, column=1)


def filas(ws, datos, desde=4):
    for i, fila in enumerate(datos, start=desde):
        for j, v in enumerate(fila, start=1):
            cel = ws.cell(row=i, column=j, value=v)
            cel.alignment = Alignment(wrap_text=True, vertical='top')
            if i % 2 == 0:
                cel.fill = PatternFill('solid', fgColor=CREMA)


def anchos(ws, mapa):
    for col, w in mapa.items():
        ws.column_dimensions[col].width = w


def si_no(ws, col, hasta=60, desde=4):
    """Desplegable TRUE/FALSE, para que nadie tenga que adivinar qué escribir."""
    dv = DataValidation(type='list', formula1='"TRUE,FALSE"', allow_blank=True)
    dv.error = 'Escribí TRUE o FALSE.'
    dv.promptTitle = 'TRUE o FALSE'
    dv.prompt = 'TRUE = sí · FALSE = no · vacío = TRUE'
    ws.add_data_validation(dv)
    dv.add(f'{col}{desde}:{col}{hasta}')


# ══════════════════════════════════════════════════════════════════════
def plantilla_base():
    wb = Workbook()

    # ── LEEME ──────────────────────────────────────────────────────────
    le = wb.active
    le.title = 'LEEME'
    manual = [
        ('Planilla del grupo', ''),
        ('', ''),
        ('Qué es', 'Esta planilla ES la configuración del sitio. Todo lo que se ve en el sitio sale de acá o de las carpetas de Drive. No hay nada cargado por formularios ni escrito en el sitio.'),
        ('Para un grupo nuevo', 'Copiá este archivo, completalo con los datos del grupo, subilo a su carpeta de Drive y pegá su link en el sitio (Configuración). No hace falta tocar nada más.'),
        ('Regla general', 'No cambies los nombres de las pestañas ni los de las columnas: el sitio los busca por ese nombre. Sí podés agregar, borrar y reordenar filas.'),
        ('Si algo está mal cargado', 'El sitio te lo dice: Configuración → «Revisión de la planilla» lista lo que no entendió o lo que falta, celda por celda.'),
        ('', ''),
        ('PESTAÑA', 'QUÉ ALIMENTA'),
        ('GRUPO', 'Nombre del grupo, descripción, objetivos y principios. Se ve en el menú, en la portada y en la sección El Grupo.'),
        ('EQUIPO', 'Quiénes coordinan o facilitan. Se ve al pie del menú lateral.'),
        ('EMPRESAS', 'Las empresas del grupo: su ficha, si están activas y cuándo no pueden presentar.'),
        ('CALENDARIO', 'Las reglas de la agenda: qué día se reúnen, cada cuánto, y cada cuántas reuniones va una ronda de novedades o una técnica.'),
        ('SIN_REUNION', 'Las semanas en que el grupo NO se reúne (vacaciones, congresos, cosecha).'),
        ('HITOS', 'La línea de tiempo del grupo.'),
        ('EJES', 'Los ejes temáticos o prioridades del año.'),
        ('EVENTOS', 'La agenda de congresos, viajes y encuentros.'),
        ('ACCESOS', 'Qué emails pueden crear una cuenta en el sitio.'),
        ('CONFIG_DRIVE', 'Opcional. Solo si el sitio no encuentra sola la carpeta de alguna empresa.'),
        ('', ''),
        ('CÓMO SE ARMA LA AGENDA', 'El sitio genera las reuniones con estas reglas, en este orden:'),
        ('1', 'Las reuniones fijadas (📌) y las marcadas como «Flexible» en el sitio no se modifican.'),
        ('2', 'Si la fecha cae en feriado y saltar_feriados = TRUE, queda «Feriado».'),
        ('3', 'Si la fecha cae dentro de un rango de SIN_REUNION, queda «Sin reunión».'),
        ('4', 'Presenta la empresa ACTIVA que hace más tiempo que no presenta, si superó semanas_entre_presentaciones y está DISPONIBLE esa fecha.'),
        ('5', 'Si ninguna empresa corresponde, la fecha se completa con ronda de novedades o reunión técnica según la proporción configurada.'),
        ('6', 'Si no hay proporciones cargadas, la fecha queda «Flexible».'),
        ('', ''),
        ('ACTIVA vs. NO_DISPONIBLE', 'Son dos cosas distintas, en la pestaña EMPRESAS:'),
        ('activa', 'FALSE = la empresa dejó de participar. Sale de la rotación del calendario, pero sigue apareciendo en Empresas y conserva su historial de reuniones.'),
        ('no_disponible', 'La empresa participa, pero hay fechas en las que no puede presentar. El calendario la saltea esos días y le asigna a otra.'),
    ]
    for i, (a, b) in enumerate(manual, start=1):
        le.cell(row=i, column=1, value=a)
        c = le.cell(row=i, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical='top')
    le['A1'].font = Font(bold=True, size=16, color=NAVY)
    for r in (3, 4, 5, 6, 20, 28):
        le.cell(row=r, column=1).font = Font(bold=True, color=NAVY)
    for r in (8,):
        for col in (1, 2):
            le.cell(row=r, column=col).font = Font(bold=True, size=11, color=NAVY)
            le.cell(row=r, column=col).fill = PatternFill('solid', fgColor=MENTA)
    for r in range(9, 19):
        le.cell(row=r, column=1).font = Font(bold=True, name='Consolas', color='2C5F6F')
    for r in (29, 30):
        le.cell(row=r, column=1).font = Font(bold=True, name='Consolas', color='2C5F6F')
    anchos(le, {'A': 24, 'B': 105})

    # ── GRUPO ──────────────────────────────────────────────────────────
    g = wb.create_sheet('GRUPO')
    titulo(g, 'GRUPO — quién es este grupo',
           'Una fila por dato. No cambies la columna «clave»: es la que lee el sitio. Escribí en «contenido».', 4)
    encabezados(g, ['clave', 'contenido', 'dónde se ve', 'obligatorio'])
    filas(g, [
        ('grupo_nombre', 'Grupo Estratégico Ejemplo', 'Menú, encabezado y pantalla de acceso', 'SÍ'),
        ('grupo_tipo', 'Directorio Colaborativo', 'Debajo del nombre, en el menú', ''),
        ('grupo_etiqueta', '', 'Etiqueta chica del menú. Dejala vacía si no hace falta', ''),
        ('facilitado_por', 'Simpleza', 'Encabezado y título de la sección de herramientas', ''),
        ('grupo_subtitulo', 'Directorio colaborativo de empresas agropecuarias.', 'Portada', ''),
        ('grupo_descripcion', 'Un espacio de trabajo colaborativo entre empresarios donde cada empresa comparte sus decisiones y recibe la mirada de sus pares.', 'Portada y sección El Grupo', 'SÍ'),
        ('grupo_inicio', '2021', 'Sección El Grupo', ''),
        ('grupo_nombre_desde', '2025', 'Sección El Grupo. Dejalo vacío si el grupo nunca cambió de nombre', ''),
        ('objetivo_general', 'Escribí acá el objetivo permanente del grupo.', 'Sección El Grupo', ''),
        ('objetivo_anio_titulo', 'Objetivo 2026', 'Título de la tarjeta del objetivo del año', ''),
        ('objetivo_2026', 'Escribí acá el objetivo de este año.', 'Sección El Grupo', ''),
        ('objetivo_2026_ampliado', 'Un párrafo más largo, si hace falta.', 'Sección El Grupo', ''),
        ('ejes_titulo', 'Ejes temáticos 2026', 'Título de la sección de ejes', ''),
        ('principio_01', 'Profundidad antes que cantidad.', 'Sección El Grupo. Agregá principio_02, principio_03…', ''),
        ('principio_02', 'Confianza, confidencialidad y participación responsable.', '', ''),
        ('frase_destacada', 'La frase que resume cómo trabaja el grupo.', 'Sección El Grupo, destacada en verde', ''),
    ])
    anchos(g, {'A': 26, 'B': 62, 'C': 48, 'D': 12})

    # ── EQUIPO ─────────────────────────────────────────────────────────
    eq = wb.create_sheet('EQUIPO')
    titulo(eq, 'EQUIPO — quiénes coordinan el grupo', 'Se ve al pie del menú lateral. Una fila por persona.', 3)
    encabezados(eq, ['nombre', 'rol', 'mostrar_en_web'])
    filas(eq, [
        ('Nombre y Apellido', 'Coordinación del grupo', 'TRUE'),
        ('Nombre y Apellido', 'Co-facilitación', 'TRUE'),
    ])
    si_no(eq, 'C')
    anchos(eq, {'A': 30, 'B': 32, 'C': 16})

    # ── EMPRESAS ───────────────────────────────────────────────────────
    em = wb.create_sheet('EMPRESAS')
    titulo(em, 'EMPRESAS — las empresas del grupo',
           'Una fila por empresa. «activa» decide si entra en la rotación del calendario; «no_disponible» dice cuándo no puede presentar. El «slug» es un nombre corto sin espacios ni tildes: no lo cambies una vez creado.', 13)
    encabezados(em, [
        'slug', 'orden', 'nombre', 'activa', 'no_disponible', 'zona', 'participantes',
        'resumen_corto', 'foco_actual', 'identidad', 'que_hace_y_como_funciona',
        'recorrido_en_el_grupo', 'mostrar_en_web',
    ])
    filas(em, [
        ('empresa-uno', 1, 'Empresa Uno S.A.', 'TRUE', 'enero, diciembre a febrero', 'Zona / provincia',
         'Nombre y Apellido, Nombre y Apellido', 'Una o dos líneas sobre la empresa.',
         'En qué está trabajando ahora.', 'Qué es la empresa, en una línea.',
         'Qué hace y cómo está organizada.', 'Por dónde pasó su proceso en el grupo.', 'TRUE'),
        ('empresa-dos', 2, 'Empresa Dos S.R.L.', 'TRUE', '1/7/2026 a 20/7/2026', 'Zona / provincia',
         'Nombre y Apellido', 'Una o dos líneas sobre la empresa.', 'En qué está trabajando ahora.',
         '', '', '', 'TRUE'),
        ('empresa-tres', 3, 'Empresa Tres', 'FALSE', '', 'Zona / provincia', 'Nombre y Apellido',
         'Dejó el grupo en 2025: queda su ficha y su historial, pero sale de la rotación.',
         '', '', '', '', 'TRUE'),
    ])
    si_no(em, 'D')
    si_no(em, 'M')
    anchos(em, {'A': 20, 'B': 7, 'C': 26, 'D': 9, 'E': 30, 'F': 22, 'G': 30,
                'H': 40, 'I': 32, 'J': 32, 'K': 40, 'L': 40, 'M': 15})
    for r in range(4, 7):
        em.row_dimensions[r].height = 62

    # ── CALENDARIO ─────────────────────────────────────────────────────
    ca = wb.create_sheet('CALENDARIO')
    titulo(ca, 'CALENDARIO — las reglas de la agenda',
           'Con esto el sitio genera las reuniones. Cambiá solo la columna «contenido».', 3)
    encabezados(ca, ['clave', 'contenido', 'qué significa'])
    filas(ca, [
        ('dia_semana', 'lunes', 'Qué día se reúne el grupo. Escribilo con letras: lunes, martes, miércoles…'),
        ('hora', '08:00', 'Solo se muestra. No cambia nada del cálculo.'),
        ('cadencia_semanas', 1, '1 = hay encuentro todas las semanas. 2 = una semana de por medio. 3 = cada tres semanas.'),
        ('saltar_feriados', 'TRUE', 'TRUE = si la fecha cae feriado no hay reunión. FALSE = se sesiona igual.'),
        ('semanas_entre_presentaciones', 26, 'Mínimo de semanas entre dos presentaciones de la misma empresa. Con 26, cada empresa presenta unas dos veces al año; las fechas que quedan libres se completan con ronda de novedades y reuniones técnicas, según las proporciones de abajo.'),
        ('proporcion_ronda_novedades', 1, 'Proporción de las fechas libres destinadas a ronda de novedades.'),
        ('proporcion_tecnica', 2, 'Proporción destinada a reunión técnica. Con 1 y 2, de cada tres fechas libres una es ronda y dos son técnicas. Con 0 y 0, quedan como Flexible.'),
    ])
    anchos(ca, {'A': 26, 'B': 16, 'C': 92})

    # ── SIN_REUNION ────────────────────────────────────────────────────
    sr = wb.create_sheet('SIN_REUNION')
    titulo(sr, 'SIN_REUNION — cuándo el grupo no se reúne',
           'Vacaciones, cosecha, congresos. El calendario deja esas fechas como «Sin reunión». Si es un solo día, dejá «hasta» vacío.', 3)
    encabezados(sr, ['desde', 'hasta', 'motivo'])
    filas(sr, [
        ('2/1/2026', '31/1/2026', 'Receso de enero'),
        ('13/7/2026', '24/7/2026', 'Vacaciones de invierno'),
    ])
    anchos(sr, {'A': 16, 'B': 16, 'C': 48})

    # ── HITOS ──────────────────────────────────────────────────────────
    hi = wb.create_sheet('HITOS')
    titulo(hi, 'HITOS — la línea de tiempo del grupo', 'Se ve en El Grupo y en Eventos e Hitos.', 6)
    encabezados(hi, ['id_hito', 'orden', 'año', 'titulo', 'descripcion', 'link', 'mostrar_en_web'])
    filas(hi, [
        ('hito-01', 1, '2021', 'Inicio del grupo', 'Qué pasó ese año.', '', 'TRUE'),
        ('hito-02', 2, '2026', 'Un hito más reciente', 'Qué pasó. El link es opcional: si está, el hito muestra un botón «Visitar ↗».', '', 'TRUE'),
    ])
    si_no(hi, 'G')
    anchos(hi, {'A': 14, 'B': 7, 'C': 9, 'D': 34, 'E': 70, 'F': 36, 'G': 16})

    # ── EJES ───────────────────────────────────────────────────────────
    ej = wb.create_sheet('EJES')
    titulo(ej, 'EJES — las prioridades del año', 'El título de la sección se cambia en GRUPO, clave ejes_titulo.', 5)
    encabezados(ej, ['id_eje', 'orden', 'titulo', 'descripcion', 'mostrar_en_web'])
    filas(ej, [
        ('eje-01', 1, 'Primer eje', 'De qué se trata.', 'TRUE'),
        ('eje-02', 2, 'Segundo eje', 'De qué se trata.', 'TRUE'),
    ])
    si_no(ej, 'E')
    anchos(ej, {'A': 14, 'B': 7, 'C': 34, 'D': 78, 'E': 16})

    # ── EVENTOS ────────────────────────────────────────────────────────
    ev = wb.create_sheet('EVENTOS')
    titulo(ev, 'EVENTOS — congresos, viajes y encuentros',
           'La fecha es texto libre: se muestra tal cual («4-5-6/8», «23 y 24/10», «Octubre 2026»). Las reuniones semanales NO van acá: las arma el calendario.', 5)
    encabezados(ev, ['fecha', 'titulo', 'descripcion', 'lugar', 'mostrar_en_web'])
    filas(ev, [
        ('4-5-6/8', 'Congreso anual', '', 'Rosario', 'TRUE'),
        ('23 y 24/10', 'Viaje del grupo', 'Recorrida por empresas de la zona.', 'Buenos Aires', 'TRUE'),
    ])
    si_no(ev, 'E')
    anchos(ev, {'A': 18, 'B': 38, 'C': 56, 'D': 24, 'E': 16})

    # ── ACCESOS ────────────────────────────────────────────────────────
    ac = wb.create_sheet('ACCESOS')
    titulo(ac, 'ACCESOS — quiénes pueden crear una cuenta',
           'Una cuenta por empresa, que la empresa comparte con su equipo. Si esta pestaña tiene filas, SOLO estos emails pueden registrarse. Igual, la coordinación autoriza cada cuenta antes del primer ingreso. Esta lista nunca se muestra en el sitio.', 4)
    encabezados(ac, ['email', 'empresa', 'nombre', 'observaciones'])
    filas(ac, [
        ('contacto@empresa-uno.com', 'Empresa Uno S.A.', 'Nombre del referente', 'La comparte con su equipo'),
    ])
    anchos(ac, {'A': 34, 'B': 26, 'C': 26, 'D': 34})

    # ── CONFIG_DRIVE ───────────────────────────────────────────────────
    cd = wb.create_sheet('CONFIG_DRIVE')
    titulo(cd, 'CONFIG_DRIVE — solo si hace falta',
           'Normalmente esta pestaña queda VACÍA: el sitio encuentra la carpeta de cada empresa buscando por nombre dentro de la carpeta de empresas. Completá una fila solo si alguna empresa no matchea (el sitio te avisa cuál).', 4)
    encabezados(cd, ['slug', 'nombre', 'id_carpeta_empresa', 'observaciones'])
    filas(cd, [
        ('empresa-uno', 'Empresa Uno S.A.', '', 'Solo si el sitio no la encontró sola'),
    ])
    anchos(cd, {'A': 20, 'B': 28, 'C': 44, 'D': 44})

    ruta = os.path.join(AQUI, 'Plantilla_Base_Grupo.xlsx')
    wb.save(ruta)
    return ruta


# ══════════════════════════════════════════════════════════════════════
def plantilla_marco():
    wb = Workbook()
    le = wb.active
    le.title = 'LEEME'
    manual = [
        ('Marco Conceptual del grupo', ''),
        ('', ''),
        ('Para qué sirve', 'Son los conceptos con los que trabaja el grupo: el vocabulario común. Cada concepto se muestra como una tarjeta en el sitio.'),
        ('Dónde se ve', 'Sitio → Recursos del grupo → tarjeta «Marco conceptual».'),
        ('Cómo se usa', 'Escribí un concepto por fila en la pestaña CONCEPTOS. Vienen seis ejemplos ya cargados: editalos, borrá los que no usen y agregá los suyos.'),
        ('', ''),
        ('COLUMNA', 'QUÉ VA'),
        ('titulo', 'OBLIGATORIO. El nombre del concepto. Ej: «Devolución». Es lo que se ve más grande en la tarjeta.'),
        ('descripcion', 'OBLIGATORIO. Qué significa, en dos o tres líneas y en criollo. Es el texto de la tarjeta.'),
        ('icono', 'Opcional. Un emoji. Si lo dejás vacío se usa 📌.'),
        ('frase_corta', 'Opcional. Tres o cuatro palabras que resumen el concepto; se ven en gris abajo del título. Si no se te ocurre, dejala vacía.'),
        ('mostrar_en_web', 'TRUE para que se vea, FALSE para esconderlo sin borrar la fila. Vacío = se muestra.'),
        ('orden', 'Opcional. Define en qué orden aparecen las tarjetas.'),
        ('', ''),
        ('Cómo conectarla', 'Subí este archivo a la carpeta del grupo en Drive, compartilo como Lector con la cuenta de servicio del sitio, y pegá su link en Configuración → Marco Conceptual.'),
    ]
    for i, (a, b) in enumerate(manual, start=1):
        le.cell(row=i, column=1, value=a)
        c = le.cell(row=i, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical='top')
    le['A1'].font = Font(bold=True, size=16, color=NAVY)
    for r in (3, 4, 5, 15):
        le.cell(row=r, column=1).font = Font(bold=True, color=NAVY)
    for col in (1, 2):
        le.cell(row=7, column=col).font = Font(bold=True, size=11, color=NAVY)
        le.cell(row=7, column=col).fill = PatternFill('solid', fgColor=MENTA)
    for r in range(8, 14):
        le.cell(row=r, column=1).font = Font(bold=True, name='Consolas', color='2C5F6F')
    anchos(le, {'A': 22, 'B': 100})

    ws = wb.create_sheet('CONCEPTOS')
    titulo(ws, 'CONCEPTOS — el vocabulario común del grupo',
           'Un concepto por fila. Los seis de abajo son ejemplos: editalos o borralos. Solo «titulo» y «descripcion» son obligatorios.', 6)
    encabezados(ws, ['titulo', 'descripcion', 'icono', 'frase_corta', 'mostrar_en_web', 'orden'])
    filas(ws, [
        ('Directorio Colaborativo',
         'Un espacio de dirección donde cada empresa presenta sus decisiones y recibe la mirada de otros empresarios. No es una consultoría: son pares que aportan desde su propia experiencia.',
         '🏛️', 'entre pares, no consultores', 'TRUE', 1),
        ('Proceso Estratégico',
         'La presentación que hace una empresa sobre un desafío concreto para recibir devoluciones del grupo. Queda registrada en la bitácora de la empresa.',
         '📋', 'presentar para decidir mejor', 'TRUE', 2),
        ('Devolución',
         'La respuesta honesta de otro empresario después de escuchar una presentación. No es un consejo ni una receta: es una perspectiva desde su propia gestión.',
         '🔄', 'escuchar y devolver', 'TRUE', 3),
        ('Línea Estratégica',
         'El área de trabajo prioritaria que una empresa define para los próximos años. Ordena las decisiones y permite seguir los avances reunión a reunión.',
         '🎯', 'de la visión a la acción', 'TRUE', 4),
        ('Brecha',
         'La distancia entre dónde está hoy la empresa y dónde quiere estar. Identificarla es el paso previo a definir en qué trabajar.',
         '📏', 'dónde estoy y dónde quiero estar', 'TRUE', 5),
        ('Bitácora',
         'El documento donde queda registrada cada reunión de la empresa en el grupo: fecha, tema y aportes. Es la memoria del proceso.',
         '📓', 'la memoria del proceso', 'TRUE', 6),
    ])
    si_no(ws, 'E')
    anchos(ws, {'A': 28, 'B': 78, 'C': 9, 'D': 32, 'E': 16, 'F': 8})
    for r in range(4, 10):
        ws.row_dimensions[r].height = 58

    ruta = os.path.join(AQUI, 'Plantilla_Marco_Conceptual.xlsx')
    wb.save(ruta)
    return ruta


if __name__ == '__main__':
    for r in (plantilla_base(), plantilla_marco()):
        print('escrita:', r)
